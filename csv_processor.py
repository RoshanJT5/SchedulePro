"""
CSV Streaming Processor for Large File Uploads
Provides memory-efficient CSV/Excel processing without pandas dependency
"""
import csv
import io
from typing import Iterator, Dict, List, Any
from openpyxl import load_workbook


def process_csv_stream(file_stream, chunk_size: int = 1000) -> Iterator[List[Dict[str, Any]]]:
    """
    Stream CSV file in chunks to avoid loading entire file into memory.
    
    Args:
        file_stream: Binary file-like object (Flask's request.files['file'].stream)
        chunk_size: Number of rows to process per chunk
        
    Yields:
        List of dictionaries representing rows in the current chunk
    """
    # Wrap binary stream in text mode for CSV reader
    text_stream = io.TextIOWrapper(file_stream, encoding='utf-8', newline='')
    reader = csv.DictReader(text_stream)
    
    chunk = []
    for i, row in enumerate(reader, 1):
        # Normalize column names to lowercase and strip whitespace
        normalized_row = {k.strip().lower(): v for k, v in row.items() if k}
        chunk.append(normalized_row)
        
        if i % chunk_size == 0:
            yield chunk
            chunk = []
    
    # Yield remaining rows
    if chunk:
        yield chunk


def process_excel_stream(file_stream, chunk_size: int = 1000) -> Iterator[List[Dict[str, Any]]]:
    """
    Stream Excel file in chunks to avoid loading entire file into memory.
    Uses openpyxl's read_only mode for memory efficiency.
    
    Args:
        file_stream: Binary file-like object (Flask's request.files['file'].stream)
        chunk_size: Number of rows to process per chunk
        
    Yields:
        List of dictionaries representing rows in the current chunk
    """
    # Load workbook in read-only mode for memory efficiency
    workbook = load_workbook(file_stream, read_only=True, data_only=True)
    sheet = workbook.active
    
    # Get headers from first row
    rows_iter = sheet.iter_rows(values_only=True)
    headers = next(rows_iter)
    
    # Normalize headers to lowercase and strip whitespace
    headers = [str(h).strip().lower() if h else f'column_{i}' for i, h in enumerate(headers)]
    
    chunk = []
    for i, row_values in enumerate(rows_iter, 1):
        # Create dictionary from headers and values
        row_dict = {}
        for header, value in zip(headers, row_values):
            # Convert None to empty string, everything else to string
            row_dict[header] = '' if value is None else str(value).strip()
        
        chunk.append(row_dict)
        
        if i % chunk_size == 0:
            yield chunk
            chunk = []
    
    # Yield remaining rows
    if chunk:
        yield chunk
    
    workbook.close()


def process_upload_stream(upload_file, chunk_size: int = 1000) -> Iterator[List[Dict[str, Any]]]:
    """
    Automatically detect file type and stream process CSV or Excel files.
    
    Args:
        upload_file: Flask FileStorage object from request.files
        chunk_size: Number of rows to process per chunk
        
    Yields:
        List of dictionaries representing rows in the current chunk
        
    Raises:
        ValueError: If file type is not supported
    """
    filename = upload_file.filename.lower()
    
    if filename.endswith('.csv'):
        yield from process_csv_stream(upload_file.stream, chunk_size)
    elif filename.endswith('.xlsx') or filename.endswith('.xls'):
        yield from process_excel_stream(upload_file.stream, chunk_size)
    else:
        raise ValueError('Unsupported file type. Upload CSV or Excel (.xlsx, .xls) files only.')


def validate_required_columns(row: Dict[str, Any], required_columns: set) -> bool:
    """
    Check if a row contains all required columns.
    
    Args:
        row: Dictionary representing a single row
        required_columns: Set of required column names (lowercase)
        
    Returns:
        True if all required columns are present, False otherwise
    """
    return required_columns.issubset(set(row.keys()))


def get_missing_columns(available_columns: set, required_columns: set) -> set:
    """
    Get the set of missing required columns.
    
    Args:
        available_columns: Set of available column names
        required_columns: Set of required column names
        
    Returns:
        Set of missing column names
    """
    return required_columns - available_columns
